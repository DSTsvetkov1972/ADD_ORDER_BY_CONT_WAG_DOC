from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook, styles
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidationList
from openpyxl.styles.borders import Border, Side

import os

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))



def make_template(ws):



    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 17   
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 19
    ws.column_dimensions["G"].width = 23
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 11
    ws.column_dimensions["M"].width = 18
    ws.column_dimensions["N"].width = 14

    # ws.row_dimensions[7].width = 36
    # ws.row_dimensions[8].width = 29
    # ws.row_dimensions[10].width = 24

    cell_khaki = styles.PatternFill(start_color='ddd9c4', fill_type='solid')

    header = {
        (7, 2): ('Наименование соисполнителя', cell_khaki, thin_border),
        (7, 3): ('№ договора с соисполнителем', cell_khaki, thin_border),
        (7, 4): ('Дата акта', cell_khaki, thin_border),
        (7, 5): ('Дата начала отчётного периода', cell_khaki, thin_border),
        (7, 6): ('Дата окончания отчётного периода', cell_khaki, thin_border),

        (10, 1): ('№', cell_khaki, thin_border),
        (10, 2): ('Дата отправки', cell_khaki, thin_border),
        (10, 3): ('Номер накладной', cell_khaki, thin_border),
        (10, 4): ('№ Контейнера', cell_khaki, thin_border),        
        (10, 5): ('№ Вагона', cell_khaki, thin_border),
        (10, 6): ('Ст. отправления', cell_khaki, thin_border),
        (10, 7): ('Ст. назначения', cell_khaki, thin_border),
        (10, 8): (' тариф груженый', cell_khaki, thin_border),
        (10, 9): ('использование пути', cell_khaki, thin_border),               
        (10, 10): ('Номер заказа', cell_khaki, thin_border),

        (10, 11): ('проверка', styles.PatternFill(), None),
        (10, 12): ('invoiceid', styles.PatternFill(), None),
        (10, 13): ('invdatecreate', styles.PatternFill(), None),               
        (10, 14): ('invfrwsubcode', styles.PatternFill(), None),
        }

    for (y, x), (cell_value, cell_fill, cell_border) in header.items():
        header_cell = ws.cell(y, x)
        header_cell.value = cell_value
        header_cell.fill = cell_fill
        header_cell.border = cell_border
        header_cell.alignment = styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
        header_cell.font = styles.Font(name='Times New Roman', size=9, bold=True, color='000000')



if __name__ == "__main__":
    wb = Workbook()
    wb.create_sheet("Отчёт", 0)
    wb.active = wb["Отчёт"]
    ws = wb["Отчёт"]
    make_template(ws)
    os.startfile("text.xlsx")
    wb.save("text.xlsx")