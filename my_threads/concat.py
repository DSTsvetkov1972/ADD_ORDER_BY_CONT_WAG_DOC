from PySide6 import QtWidgets, QtCore
from colorama import Fore
from datetime import datetime
import global_vars 
import os, random
import pandas as pd
from time import sleep
from my_threads.functions import check_files_modified
from openpyxl import load_workbook, styles
from openpyxl.utils.cell import get_column_letter

class ConcatThread(QtCore.QThread):
 
    mysignal = QtCore.Signal(str)


    def __init__ (self, parent=None):
        QtCore.QThread.__init__(self, parent) 
        self.message_title = "Объединение"
        
        
    def clean_folder_marked(self, project_folder):

        errors_list = []


        source_files = list(os.walk(os.path.join(global_vars.project_folder,'.Исходники')))[0][2]
        marked_files = [file for file in list(os.walk(os.path.join(global_vars.project_folder,'.Размеченные')))[0][2] if file[0] != '~']
     
        for file in marked_files:
            sleep(0.0001)
            self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                               f"Проверяем наличие файла {file} из ./Размеченные в .Исходники/")             
            if file[3:] not in source_files: # file[3:] чтобы откусить приставку md_ в начале
                try:
                    os.remove(os.path.join(global_vars.project_folder, '.Размеченные', file))
                except PermissionError:
                    errors_list.append("Книга {file} есть в папке .Размеченные/,\
                            но ей нет соответствия в папке ./Исходники.\nНе можем удалить эту книгу из ./Размеченные,\
                                       потому что она открыта в Эксель!")

        return ("\n" + ">" + "\n").join(errors_list)


    def concat_dfs(self, project_folder):
        global_vars.ui.info_label.setStyleSheet('color: blue')
        #project_folder = os.path.join(r'C:\Users\TsvetkovDS\Documents\Оперативная папка\.Тест')
        random_suffix = random.randrange(0,1000000)
        file_field_name = f"file_({random_suffix})"
        sheet_field_name = f"sheet_({random_suffix})"        
        source_row_field_name = f"source_row_({random_suffix})"

        marked_folder = os.path.join(project_folder, r".Размеченные")
        files = [file for file in list(os.walk(os.path.join(project_folder, '.Размеченные')))[0][2] if file[0] != "~"]
        columns_info_df = pd.read_excel(os.path.join(project_folder,'columns.xlsx'))
        dfs_to_concat = [pd.DataFrame()]

        for file_number, file in enumerate(files, 1):
            with pd.ExcelFile(os.path.join(marked_folder,file)) as xlsx_file:
                sheets = xlsx_file.sheet_names
                    
            for sheet_number, sheet in enumerate(sheets, 1):
                # sleep(0.0001)
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {file_number} из {len(files)} лист {sheet_number} из {len(sheets)}. " 
                                   f"Подготавливаем к объединению {file} лист {sheet}")  
                file_info = columns_info_df[(columns_info_df['file'] == file) &
                                            (columns_info_df['sheet'] == sheet) &
                                            (columns_info_df['Ошибки маркировки'] == 'ok')]
                
                if not file_info.empty:
                    s = int(file_info['s'].iloc[0])-1
                    f = int(file_info['f'].iloc[0])
                    file_df = pd.read_excel(os.path.join(marked_folder,file), sheet_name=sheet, header=None).iloc[:,2:]

        
                    file_df_without_ffill = file_df.iloc[s:f]
                    file_df_without_ffill.columns = file_df.iloc[0]
                    column_names_without_ffill = [column_name for column_name in file_df.iloc[0] if pd.notna(column_name)]
                    file_df_without_ffill = file_df_without_ffill[column_names_without_ffill]

                    # file_df_with_ffill = file_df.fillna(method='ffill')     
                    file_df_with_ffill = file_df.ffill()                       
                    file_df_with_ffill = file_df_with_ffill.iloc[s:f]
                    file_df_with_ffill.columns = file_df.iloc[1]
                    column_names_with_ffill = [column_name for column_name in file_df.iloc[1] if pd.notna(column_name)]
                    file_df_with_ffill = file_df_with_ffill[column_names_with_ffill]           

                    file_df = pd.concat([file_df_without_ffill, file_df_with_ffill], axis=1)

                    file_df[file_field_name] = file
                    file_df[sheet_field_name] = sheet
                    file_df[source_row_field_name] = file_df.index + 1

                    dfs_to_concat.append(file_df)


        if dfs_to_concat:
            self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                               f"Создаём итоговую таблицу") 
            # sleep(0.0001) 
            result_df = pd.concat(dfs_to_concat)
         
            result_df_columns = list(result_df.columns)
            
            if (file_field_name in result_df_columns and
                sheet_field_name in result_df_columns and 
                source_row_field_name in result_df_columns):
                
                result_df_columns.remove(file_field_name)
                result_df_columns.remove(sheet_field_name)
                result_df_columns.remove(source_row_field_name)
                result_df = result_df[result_df_columns + [file_field_name, sheet_field_name, source_row_field_name]]
                self.result_df_len = len(result_df)
            

            if  self.result_df_len < 1048576: 
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Записываем результат в result.xlsx'")   
                sleep(0.0001)   
                res_file_name = os.path.join(project_folder,'result.xlsx')           
                result_df.to_excel(res_file_name, index=False)

                # задаём ширину столбцов по размеру заголовка
                wb = load_workbook(res_file_name)
                ws = wb.active

                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Подгоняем ширину столбцов под длины заголовков")

                for n, column in enumerate(list(result_df.columns), 1):
                    ws.column_dimensions[get_column_letter(n)].width = len(str(column))*1.1 + 5

                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Замораживаем строку заголовков")
                ws.auto_filter.ref = ws.dimensions    
                
                ws.freeze_panes = ws.cell(column=1, row=2)

                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Задаём цвет строки заголовков")
                for col, column in enumerate(list(result_df.columns), start=1):
                    cell = ws.cell(column=col, row = 1)
                    # cell.fill = styles.PatternFill(start_color='FFFFC7CE', fill_type='solid')
                    cell.fill = styles.PatternFill(start_color='FDE9D9', fill_type='solid')
                    cell.font = styles.Font(color='974706', bold=True)
                    cell.alignment = styles.Alignment(wrap_text=True,
                                                    vertical='top',
                                                    horizontal='center') 
                    #cell.style.alignment.wrap_text=True

                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                    f"Сохраняем файл")    
                wb.save(res_file_name)
                os.startfile(os.path.join(project_folder,'result.xlsx'))
            else:
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Записываем результат в result.csv")                 
                result_df.to_csv(os.path.join(project_folder,'result.csv', sep = '\t'), index=False)
                os.startfile(os.path.join(project_folder,'result.csv'))                


    def on_signal(self,mysignal):
        global_vars.ui.info_label.setStyleSheet('color: blue')            
        global_vars.ui.info_label.setText(mysignal)


    def run(self): 
        self.error_message = ""
        self.warning_message = ""
        self.result_df_len = 0

        self.is_src_files_modifyed = check_files_modified('.Исходники')
        self.is_md_files_modifyed = check_files_modified('.Размеченные')

        if os.path.exists(os.path.join(global_vars.project_folder, "~$result.xlsx")):
            global_vars.ui.info_label.setStyleSheet('color: red') 
            self.error_message = 'Закройте файл result.xlsx и снова нажмите "Объединить"'
            global_vars.ui.info_label.setText(self.error_message)
            os.startfile(os.path.join(global_vars.project_folder, "result.xlsx"))
            return
        
        if os.path.exists(os.path.join(global_vars.project_folder, "~$result.csv")):
            global_vars.ui.info_label.setStyleSheet('color: red')
            self.error_message = 'Файл result.csv занят другим приложением и не может быть перезаприсан!'
            global_vars.ui.info_label.setText(self.error_message)
            # os.startfile(os.path.join(global_vars.project_folder, "result.csv"))
            return

        if self.is_src_files_modifyed:
            global_vars.ui.info_label.setStyleSheet('color: red')
            self.error_message = ('В папку .Исходники были добавлены новые файлы или\n'
                                  'некоторые файлы в ней были пересохранены или удалены.\n'
                                  'Перезапуститет обработку!')
            # global_vars.ui.info_label.setText(self.error_message)
            return

        if self.is_md_files_modifyed:
            global_vars.ui.info_label.setStyleSheet('color: red')
            self.error_message = ('Файлы в папке .Размеченные были изменены.\n'
                                  'Перезапустите обработку!')  
            # global_vars.ui.info_label.setText(self.error_message)
            return  
              
        if os.path.exists(os.path.join(global_vars.project_folder, "result.xlsx")):
            try:
                os.remove(os.path.join(global_vars.project_folder, "result.xlsx"))
            except PermissionError:
                global_vars.ui.info_label.setStyleSheet('color: red') 
                self.error_message = 'Закройте файл result.xlsx и снова нажмите "Объединить"'                
                #global_vars.ui.info_label.setText(self.error_message)
                os.startfile(os.path.join(global_vars.project_folder, "result.xlsx"))                
                return

        if os.path.exists(os.path.join(global_vars.project_folder, "result.xlsx")):
            try:
                os.remove(os.path.join(global_vars.project_folder, "result.xlsx"))
            except PermissionError:
                global_vars.ui.info_label.setStyleSheet('color: red') 
                self.error_message = 'Файл result.csv занят другим приложением и не может быть перезаписан!'
                #global_vars.ui.info_label.setText(self.error_message)
                # os.startfile(os.path.join(global_vars.project_folder, "result.#sv"))
                return

        self.error_message = ""
        self.warning_message = ""        
        self.error_message = self.clean_folder_marked(global_vars.project_folder)
        if not self.error_message:  
            self.concat_dfs(global_vars.project_folder)


    def on_clicked(self):     
        self.start() # Запускаем поток  
    

    def on_started(self): # Вызывается при запуске потока
        global_vars.ui.pushButtonChooseProjectFolder.setEnabled(False)  
        global_vars.ui.pushButtonXLStoXLSX.setEnabled(False)      
        global_vars.ui.pushButtonProcessing.setEnabled(False)
        global_vars.ui.pushButtonOpenChoosedFiles.setEnabled(False)        
        global_vars.ui.pushButtonOpenChoosedMDFiles.setEnabled(False)
        global_vars.ui.pushButtonDelChoosedMDFiles.setEnabled(False)           
        global_vars.ui.pushButtonConcat.setEnabled(False)        
        global_vars.ui.pushButtonMakeFiles.setEnabled(False)    

    def on_finished(self): # Вызывается при завершении потока
        global_vars.ui.pushButtonChooseProjectFolder.setEnabled(True)
        global_vars.ui.pushButtonXLStoXLSX.setEnabled(True)        
        global_vars.ui.pushButtonProcessing.setEnabled(True)
        global_vars.ui.pushButtonOpenChoosedFiles.setEnabled(True)
        global_vars.ui.pushButtonOpenChoosedMDFiles.setEnabled(True)
        global_vars.ui.pushButtonDelChoosedMDFiles.setEnabled(True)

        if self.is_src_files_modifyed or self.is_md_files_modifyed:
            global_vars.ui.pushButtonConcat.setEnabled(False)
            global_vars.ui.pushButtonMakeFiles.setEnabled(False)
        else:
            global_vars.ui.pushButtonConcat.setEnabled(True) 
            global_vars.ui.pushButtonMakeFiles.setEnabled(True)
         

        if self.error_message:
            global_vars.ui.info_label.setStyleSheet('color: red')
            global_vars.ui.info_label.setText(self.error_message.replace('\n',' '))
            QtWidgets.QMessageBox.critical(None,
                                           self.message_title,
                                           self.error_message,
                                           buttons=QtWidgets.QMessageBox.StandardButton.Ok) 
        elif self.warning_message:
            global_vars.ui.info_label.setStyleSheet('color: red')
            global_vars.ui.info_label.setText(self.warning_message.replace('\n',' '))
            QtWidgets.QMessageBox.warning(None,
                                           self.message_title,
                                           self.warning_message,
                                           buttons=QtWidgets.QMessageBox.StandardButton.Ok)
        else:
            global_vars.ui.info_label.setStyleSheet('color: green')
            if self.result_df_len < 1048576:          
                global_vars.ui.info_label.setText(f'Результат содержит {self.result_df_len} строк и загружен файл result.xlsx')
            else:
                global_vars.ui.info_label.setText(f'Результат содержит {self.result_df_len} строк и загружен файл result.csv')
