from PySide6 import QtWidgets, QtCore
from colorama import Fore
from datetime import datetime
import global_vars 
import os
import shutil
import pandas as pd
from time import sleep
from my_threads.functions import check_files_modified
#from openpyxl import load_workbook, 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.borders import Border, Side
from openpyxl import styles
#from openpyxl.utils.cell import get_column_letter
from my_functions.sql import sql
from my_functions.dwh import get_df_of_click
from my_functions.for_make_files import make_template
import pyperclip

class MakeFilesThread(QtCore.QThread):
 
    mysignal = QtCore.Signal(str)

    def on_signal(self,mysignal):
        global_vars.ui.info_label.setStyleSheet('color: blue')
        print("vs nen")
        print(str(mysignal))       
        global_vars.ui.info_label.setText(mysignal)


    def __init__ (self, parent=None):
        QtCore.QThread.__init__(self, parent) 
        self.message_title = "Делаем файлы"
        
        
    def clean_folder_marked(self, project_folder):

        errors_list = []

        source_files = list(os.walk(os.path.join(global_vars.project_folder,'.Исходники')))[0][2]
        marked_files = [file for file in list(os.walk(os.path.join(global_vars.project_folder,'.Размеченные')))[0][2] if file[0] != '~']
     
        for file in marked_files:
            sleep(0.0001)
            self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                               f"Проверяем наличие файла {file} из ./Размеченные в .Исходники/")             
            if file[3:] not in source_files: # file[3:] чтобы откусить приставку md_ в начале
                try:
                    os.remove(os.path.join(global_vars.project_folder, '.Размеченные', file))
                except PermissionError:
                    errors_list.append("Книга {file} есть в папке .Размеченные/,\
                            но ей нет соответствия в папке ./Исходники.\nНе можем удалить эту книгу из ./Размеченные,\
                                       потому что она открыта в Эксель!")

        return ("\n" + ">" + "\n").join(errors_list)




    def concat_dfs(self, project_folder):
        global_vars.ui.info_label.setStyleSheet('color: blue')
        #project_folder = os.path.join(r'C:\Users\TsvetkovDS\Documents\Оперативная папка\.Тест')
        #random_suffix = random.randrange(0,1000000)
        #file_field_name = f"file_({random_suffix})"
        #sheet_field_name = f"sheet_({random_suffix})"        
        #source_row_field_name = f"source_row_({random_suffix})"

        file_field_name = "source_file"
        sheet_field_name = "sorce_sheet"        
        source_row_field_name = "source_row"        

        marked_folder = os.path.join(project_folder, r".Размеченные")
        files = [file for file in list(os.walk(os.path.join(project_folder, '.Размеченные')))[0][2] if file[0] != "~"]
        columns_info_df = pd.read_excel(os.path.join(project_folder,'columns.xlsx'))
        dfs_to_concat = [pd.DataFrame()]

        for file_number, file in enumerate(files, 1):
            with pd.ExcelFile(os.path.join(marked_folder,file)) as xlsx_file:
                sheets = xlsx_file.sheet_names
                    
            for sheet_number, sheet in enumerate(sheets, 1):
                # sleep(0.0001)
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {file_number} из {len(files)} лист {sheet_number} из {len(sheets)}. " 
                                   f"Подготавливаем к объединению {file} лист {sheet}")  
                file_info = columns_info_df[(columns_info_df['file'] == file) &
                                            (columns_info_df['sheet'] == sheet) &
                                            (columns_info_df['Ошибки маркировки'] == 'ok')]
                
                if not file_info.empty:
                    s = int(file_info['s'].iloc[0]) - 1
                    f = int(file_info['f'].iloc[0])
                    file_df = pd.read_excel(os.path.join(marked_folder,file), sheet_name=sheet, header=None).iloc[:,2:]

                    file_df_without_ffill = file_df.iloc[s:f]
                    file_df_without_ffill.columns = file_df.iloc[0]
                    column_names_without_ffill = [column_name for column_name in file_df.iloc[0] if pd.notna(column_name)]
                    file_df_without_ffill = file_df_without_ffill[column_names_without_ffill]

                    # file_df_with_ffill = file_df.fillna(method='ffill')     
                    file_df_with_ffill = file_df.ffill()                       
                    file_df_with_ffill = file_df_with_ffill.iloc[s:f]
                    file_df_with_ffill.columns = file_df.iloc[1]
                    column_names_with_ffill = [column_name for column_name in file_df.iloc[1] if pd.notna(column_name)]
                    file_df_with_ffill = file_df_with_ffill[column_names_with_ffill]           

                    file_df = pd.concat([file_df_without_ffill, file_df_with_ffill], axis=1)

                    file_df[file_field_name] = file
                    file_df[sheet_field_name] = sheet
                    file_df[source_row_field_name] = file_df.index + 1
   
                    dfs_to_concat.append(file_df)


        self.mysignal.emit(
            f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
            f"Создаём список вагонов, накладных, контейнеров для загрузки в DWH"
            ) 


        result_df = pd.concat(dfs_to_concat)

        
        result_df_columns = list(result_df.columns)
        
        if (file_field_name in result_df_columns and
            sheet_field_name in result_df_columns and 
            source_row_field_name in result_df_columns):
            
            result_df_columns.remove(file_field_name)
            result_df_columns.remove(sheet_field_name)
            result_df_columns.remove(source_row_field_name)
            result_df = result_df[result_df_columns + [file_field_name, sheet_field_name, source_row_field_name]]
            self.result_df_len = len(result_df)
        
        if result_df.empty:
            global_vars.ui.info_label.setStyleSheet('color: red') 
            self.error_message = 'Нет корректных размеченных файлов!'                
            return

        print("Датафреймы слиты в один для загрузки в DWH")
        result_df['№ Вагона'] = result_df['№ Вагона'].apply(str)
        result_df['Номер накладной'] = result_df['Номер накладной'].apply(str)
        result_df['№ Контейнера'] = result_df['№ Контейнера'].apply(str)

        result_df['Сцеп'] = (
            result_df['№ Вагона'] + '|' +
            result_df['Номер накладной'] + '|' +
            result_df['№ Контейнера']
        )
    
        
        scep_series = result_df['Сцеп']
        #print(scep_series)
        scep_series.drop_duplicates(inplace=True)
        scep_str = '\n'.join(scep_series)
        #pyperclip.copy(scep_str)
        sql_str = sql(scep_str)
        #pyperclip.copy(sql_str)
        #print("в буфере")

        self.mysignal.emit(
            f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
            f"Получаем данные из DWH"
            )

        
        sql_res_df = get_df_of_click(sql_str)
        


        sql_res_df = sql_res_df.fillna("")

        result_df = result_df.merge(sql_res_df, how='left', on=['Сцеп'])
        result_df.rename(columns={'Номер_заказа': 'Номер заказа'}, inplace=True)
        
        file_sheet_df = result_df.groupby(['source_file', 'sorce_sheet'])
        file_sheet_df = pd.DataFrame(file_sheet_df)
        file_sheet_df = file_sheet_df.apply(lambda x: x)
        for file_sheet_tuple in file_sheet_df.itertuples():
            file = file_sheet_tuple[1][0]
            sheet = file_sheet_tuple[1][1]

            file_1s = f'Форма отчета исполнителя МЛТ груженный, порожний_{file}_{sheet}.xlsx'                        

            self.mysignal.emit(
                f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                f"Заполняем { file_1s }"
            )
            
            #processing_file = os.path.join(
            #    project_folder,
            #    '.Обработка',
            #    file_1s 
            #    )
            
            ready_file = os.path.join(
                project_folder,
                '.Файлы для 1-С',
                file_1s 
                )
            
            
            df_to_1s = file_sheet_tuple[2][[
                '№', 'Дата отправки', 'Номер накладной', '№ Контейнера', '№ Вагона',
                'Ст. отправления', 'Ст. назначения', ' тариф груженый', 'использование пути', 'Номер заказа',
                'проверка', 'invoiceid' ,'invdatecreate', 'invfrwsubcode'
            ]]
            
            """
            shutil.copy(
                os.path.join(
                    project_folder,
                    'Форма отчета исполнителя МЛТ груженный, порожний.xlsx'
                    ),
                processing_file
            )

            wb = load_workbook(processing_file)
            """

            wb = Workbook()
            wb.create_sheet("Отчёт", 0)
            wb.active = wb["Отчёт"]
            ws = wb["Отчёт"]

            make_template(ws)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
                )

            
            rows = dataframe_to_rows(df_to_1s)
            rows = [row[1:] for row in list(rows)[2:]]
            for r_idx, row in enumerate(rows, 11):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    ws.cell(row=r_idx, column=c_idx).font = styles.Font(name='Times New Roman', size=9, bold=False, color='000000')
                    ws.cell(row=r_idx, column=c_idx).alignment = styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
                    if value and c_idx==11:
                        ws.cell(row=r_idx, column=c_idx).fill = styles.PatternFill(start_color='ff0000', fill_type='solid')
                        ws.cell(row=10, column=11).fill = styles.PatternFill(start_color='ff0000', fill_type='solid')
                        filters = ws.auto_filter
                        filters.ref = "K10"
                    if c_idx<11:
                        # print(r_idx, c_idx)
                        ws.cell(row=r_idx, column=c_idx).border = thin_border

            # print(file_sheet_tuple[2]['Отчет к акту выполненных работ №'])
            # input()        

            ws['A1'] =  file_sheet_tuple[2]['Отчет к акту выполненных работ №'].iloc[0]
            ws['A1'].font = styles.Font(name='Times New Roman', size=11, bold=False, color='000000')
            ws['A1'].alignment = styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
            ws.merge_cells("A1:H1")
            ws['A2'] =  file_sheet_tuple[2]['Между'].iloc[0]
            ws['A2'].font = styles.Font(name='Times New Roman', size=11, bold=False, color='000000')
            ws['A2'].alignment = styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
            ws.merge_cells("A2:H2")

            ws['H4'] =  file_sheet_tuple[2]['Дата составления'].iloc[0]
            ws['H4'].font = styles.Font(name='Times New Roman', size=9, italic=True, bold=False, color='000000')
            ws['H4'].alignment = styles.Alignment(wrap_text=False, horizontal="right", vertical="center")

            #for cell in ['A1','A2','H4']:
                #ws[cell].alignment = styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
                #   ws[cell].font = styles.Font(name='Times New Roman', size=9, bold=False, color='000000')
                #ws.row_dimensions[8].width = 29    

            
            ws['B8'] = file_sheet_tuple[2]['Наименование соисполнителя'].iloc[0]
            ws['C8'] = file_sheet_tuple[2]['№ договора с соисполнителем'].iloc[0]
            ws['D8'] = file_sheet_tuple[2]['Дата акта'].iloc[0]
            ws['E8'] = file_sheet_tuple[2]['Дата начала отчётного периода'].iloc[0]
            ws['F8'] = file_sheet_tuple[2]['Дата окончания отчётного периода'].iloc[0]

            for cell in ['B8','C8','D8','E8','F8']:
                ws[cell].border = thin_border
                ws[cell].alignment = styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
                ws[cell].font = styles.Font(name='Times New Roman', size=9, bold=False, color='000000')
                #ws.row_dimensions[8].width = 29                                                                         

            wb.save(ready_file)        
            # os.remove(processing_file)




    def run(self): 
        self.error_message = ""
        self.warning_message = ""
        self.result_df_len = 0

        self.is_src_files_modifyed = check_files_modified('.Исходники')
        self.is_md_files_modifyed = check_files_modified('.Размеченные')

        """
        if os.path.exists(os.path.join(global_vars.project_folder, "~$result.xlsx")):
            global_vars.ui.info_label.setStyleSheet('color: red') 
            self.error_message = 'Закройте файл result.xlsx и снова нажмите "Объединить"'                
            global_vars.ui.info_label.setText(self.error_message)
            os.startfile(os.path.join(global_vars.project_folder, "result.xlsx"))                   
            return
        
        if os.path.exists(os.path.join(global_vars.project_folder, "~$result.csv")):
            global_vars.ui.info_label.setStyleSheet('color: red') 
            self.error_message = 'Файл result.csv занят другим приложением и не может быть перезаприсан!'                
            global_vars.ui.info_label.setText(self.error_message)
            # os.startfile(os.path.join(global_vars.project_folder, "result.csv"))                   
            return
        """

        if self.is_src_files_modifyed:
            global_vars.ui.info_label.setStyleSheet('color: red')
            self.error_message = ('В папку .Исходники были добавлены новые файлы или\n'
                                  'некоторые файлы в ней были пересохранены или удалены.\n'
                                  'Перезапуститет обработку!')
            # global_vars.ui.info_label.setText(self.error_message)
            return

        if self.is_md_files_modifyed:
            global_vars.ui.info_label.setStyleSheet('color: red')
            self.error_message = ('Файлы в папке .Размеченные были изменены.\n'
                                  'Перезапустите обработку!')  
            # global_vars.ui.info_label.setText(self.error_message)
            return  
        """      
        if os.path.exists(os.path.join(global_vars.project_folder, "result.xlsx")):
            try:
                os.remove(os.path.join(global_vars.project_folder, "result.xlsx"))
            except PermissionError:
                global_vars.ui.info_label.setStyleSheet('color: red') 
                self.error_message = 'Закройте файл result.xlsx и снова нажмите "Объединить"'                
                #global_vars.ui.info_label.setText(self.error_message)
                os.startfile(os.path.join(global_vars.project_folder, "result.xlsx"))                
                return
        

        if os.path.exists(os.path.join(global_vars.project_folder, "result.xlsx")):
            try:
                os.remove(os.path.join(global_vars.project_folder, "result.xlsx"))
            except PermissionError:
                global_vars.ui.info_label.setStyleSheet('color: red') 
                self.error_message = 'Файл result.csv занят другим приложением и не может быть перезаписан!'                
                #global_vars.ui.info_label.setText(self.error_message)
                # os.startfile(os.path.join(global_vars.project_folder, "result.#sv"))                   
                return
        """

        self.error_message = ""
        self.warning_message = ""        
        self.error_message = self.clean_folder_marked(global_vars.project_folder)
        if not self.error_message:
            try:
                shutil.rmtree(
                    os.path.join(
                        global_vars.project_folder,
                        '.Файлы для 1-С'
                    )
                )
            except PermissionError:
                global_vars.ui.info_label.setStyleSheet('color: red') 
                self.error_message = 'Некоторые файлы из папки ".Файлы для 1-С" открыты на рабочем столе. Закройте их и снова запустите создание файлов!'                
                return
            os.mkdir(
                os.path.join(
                    global_vars.project_folder,
                    '.Файлы для 1-С'
                )
            )
            self.concat_dfs(global_vars.project_folder)


    def on_clicked(self):     
        self.start() # Запускаем поток  
    

    def on_started(self): # Вызывается при запуске потока
        global_vars.ui.pushButtonChooseProjectFolder.setEnabled(False)        
        global_vars.ui.pushButtonProcessing.setEnabled(False)
        global_vars.ui.pushButtonOpenChoosedFiles.setEnabled(False)        
        global_vars.ui.pushButtonOpenChoosedMDFiles.setEnabled(False)
        global_vars.ui.pushButtonDelChoosedMDFiles.setEnabled(False)           
        global_vars.ui.pushButtonConcat.setEnabled(False)        
        global_vars.ui.pushButtonMakeFiles.setEnabled(False)

    def on_finished(self): # Вызывается при завершении потока
        global_vars.ui.pushButtonChooseProjectFolder.setEnabled(True)  
        global_vars.ui.pushButtonProcessing.setEnabled(True)
        global_vars.ui.pushButtonOpenChoosedFiles.setEnabled(True)        
        global_vars.ui.pushButtonOpenChoosedMDFiles.setEnabled(True)
        global_vars.ui.pushButtonDelChoosedMDFiles.setEnabled(True)           

        if self.is_src_files_modifyed or self.is_md_files_modifyed:
            global_vars.ui.pushButtonConcat.setEnabled(False)
            global_vars.ui.pushButtonMakeFiles.setEnabled(False)                 
        else:
            global_vars.ui.pushButtonConcat.setEnabled(True)
            global_vars.ui.pushButtonMakeFiles.setEnabled(True)             

         

        if self.error_message:
            global_vars.ui.info_label.setStyleSheet('color: red')             
            global_vars.ui.info_label.setText(self.error_message.replace('\n',' '))
            QtWidgets.QMessageBox.critical(None,
                                           self.message_title,
                                           self.error_message,
                                           buttons=QtWidgets.QMessageBox.StandardButton.Ok) 
        elif self.warning_message:
            global_vars.ui.info_label.setStyleSheet('color: red')             
            global_vars.ui.info_label.setText(self.warning_message.replace('\n',' '))
            QtWidgets.QMessageBox.warning(None,
                                           self.message_title,
                                           self.warning_message,
                                           buttons=QtWidgets.QMessageBox.StandardButton.Ok)             
        else:
            global_vars.ui.info_label.setStyleSheet('color: green')             
            global_vars.ui.info_label.setText('Файлы для загрузки в 1-С подготовленны!')
            os.startfile(os.path.join(global_vars.project_folder, '.Файлы для 1-С'))