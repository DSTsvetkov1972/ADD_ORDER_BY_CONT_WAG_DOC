from PySide6 import QtWidgets, QtCore
from colorama import Fore
import global_vars 
import os, shutil
import pandas as pd

from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook, styles
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidationList

from datetime import datetime
from time import sleep
from my_threads.functions import value_searcher, marking_checker
from my_threads.functions import init_project, refresh_files_info, clean_process_folder, check_files_modified

horizontal_offset = 2
vertical_offset = 2
sep_cell_style = styles.PatternFill(start_color='FFFFC7CE', fill_type='solid')
no_fill = styles.PatternFill(fill_type=None)
side = styles.Side(border_style=None)
no_border = styles.borders.Border(
    left=side,
    right=side,
    top=side,
    bottom=side,
)


class ProcessingThread(QtCore.QThread):
 
    mysignal = QtCore.Signal(str)


    def __init__ (self, parent=None):
        QtCore.QThread.__init__(self, parent)

           
    def check_src_files_available(self):
        """
        Перед началом обработки проверяем чтобы не было 
        открытых исходных файлов
        """

        self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                            f"проверяем содержимое папки .Исходники")

        src_files = list(os.walk(os.path.join(global_vars.project_folder,'.Исходники')))[0][2]
        self.err_list = [(f'md_{file}', 'Файл из папки .Исходники открыт на рабочем столе. Его нужно закрыть!') for file in src_files if os.path.exists(os.path.join(global_vars.project_folder, '.Исходники', f'~${file}'))]

        if self.err_list:
            self.error_message = ("Некоторых файлы из папки .Исходники,\n"
                                  "открыты на рабочем столе.")
            



    def clean_md_folder(self):
        """
        Удаляет файл из папки .Размеченные, если его нет в папке .Исходники.
        Если файл не удаётся удалить, т.к. он открыть в другой программе,
        информация о нем добавляется в список ошибок.
        """

        self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                            f"проверяем содержимое папки .Размеченные")

        source_files = list(os.walk(os.path.join(global_vars.project_folder,'.Исходники')))[0][2]
        marked_files = [file for file in list(os.walk(os.path.join(global_vars.project_folder,'.Размеченные')))[0][2] if file[0:2] != '~$']
        
        for md_file in marked_files:
                        
            if md_file[3:] not in source_files: # file[3:] чтобы откусить приставку md_ в начале
                try:
                    os.remove(os.path.join(global_vars.project_folder, '.Размеченные', md_file))
                except PermissionError:
                    # os.startfile(os.path.join(global_vars.project_folder, '.Размеченные', md_file))
                    self.error_message = ("Некоторых файлов нет в папке .Исходники,\n"
                                            "но соответствующие md-файлы не могут быть удалены,\n"
                                            "т.к. открыты в другой программе")
                    self.err_list.append((md_file, 'файла нет в .Исходники, но md файл не может быть удален, т.к. открыт в другой программе'))


    
    def check_md_book_sheet(self, wb):
        """
        Прверяем можем ли мы сдвинуть содержимое листа вниз и вправо
        или размер данных не позволяет это сделать
        """
        sheets_exceeding_dict = {}
        for sheet_number, sheet in enumerate(wb.sheetnames, 1):
            sheets_exceeding_dict[sheet] = ''
            ws = wb[sheet]

            # проверяем на размер листа
            ws_max_column = ws.max_column
            ws_max_row = ws.max_row

            ws_max_rows = 1048574-vertical_offset-1
            ws_mas_columns = 16384-horizontal_offset-1

            if ws_max_column > ws_mas_columns:
                sheets_exceeding_dict[sheet] += f' колонок больше {ws_mas_columns}'

            if ws_max_row > ws_max_rows:
                sheets_exceeding_dict[sheet] += f'строк больше {ws_max_rows}'

            if not sheets_exceeding_dict[sheet]:
                sheets_exceeding_dict.pop(sheet)

        return sheets_exceeding_dict


    def check_md_book(self, md_file, prc_file):
        """
        Почемуто некоторые Экселевские файлы не могут быть открыты в openpyxl если их
        не пересохранить.
        Проверяем файл на эту ошибку.
        """

        # пытаемся удалить файл-костыль если он образовался на предыдущем шаге
        if os.path.exists(os.path.join(global_vars.project_folder, '.Обработка', '~~~if_accident.xlsx')):
            try:
                os.remove(os.path.join(global_vars.project_folder, '.Обработка', '~~~if_accident.xlsx'))
            except:
                print(Fore.RED, 'Не удалось удалить ~~~if_accident.xlsx', Fore.RESET)


        try:

            wb = load_workbook(os.path.join(global_vars.project_folder, '.Обработка', prc_file),
                               data_only=True)

            size_check = self.check_md_book_sheet(wb) # если удалось книгу открыть, проверяем на размер данных на листах

            if size_check:
                self.error_message = "Некоторые файлы в папке .Исходники не могут быть обработаны."
                self.err_list.append((md_file, str(size_check)))
                os.remove(os.path.join(global_vars.project_folder, '.Обработка', prc_file))
                return False
            else:
                return wb
            #return wb
        except Exception:
            print(f'Авария 1 {prc_file}')
            self.error_message = "Некоторые файлы в папке .Исходники не могут быть обработаны."
            self.err_list.append((md_file, 'Возможно файл повреждён. Попробуйте пересохранить исходный файл.'))


        # костыль, чтобы освободить файл, который не удалось загрузить 
        # в предыдущем try except блоке
        try:
            print('Авария 2')
            # shutil.copy(os.path.join(global_vars.project_folder, '.Размеченные', file), os.path.join(global_vars.project_folder, '.Аварийные', file))
            shutil.copy(os.path.join(global_vars.project_folder, '.Обработка', prc_file),
                        os.path.join(global_vars.project_folder, '.Обработка', '~~~if_accident.xlsx'))
            print('Авария 3')
            # wb = load_workbook(os.path.join(global_vars.project_folder, '.Аварийные', file), data_only=True)
            wb = load_workbook(os.path.join(global_vars.project_folder, '.Обработка', '~~~if_accident.xlsx'), data_only=True)
        except:
            pass

        try:
            print('Авария 4')
            os.remove(os.path.join(global_vars.project_folder, '.Обработка', prc_file))
            print('Авария 5')
        except:
            print('Авария 666')

        return False


    def premarker(self, project_folder):
        global_vars.ui.info_label.setStyleSheet('color: blue')

        source_files = list(os.walk(os.path.join(global_vars.project_folder,'.Исходники')))[0][2]
        prc_files = list(os.walk(os.path.join(global_vars.project_folder,'.Обработка')))[0][2]
        md_files = list(os.walk(os.path.join(global_vars.project_folder,'.Размеченные')))[0][2]
        
        for source_file_number, source_file in enumerate(source_files, 1):
            print(Fore.YELLOW ,source_file, Fore.RESET)
            md_file = 'md_' + source_file

            if md_file in md_files: 
                continue

            self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                    f"Книга {source_file_number} из {len(source_files)}. Создаём размеченную книгу для {source_file}")

            prc_file = 'prc_' + source_file

            # бывало что prc файл блокировался при проверке и его не 
            # получалось удалить
            # на этот случай придуман костыль создающий prc файл с другим 
            # именем
            while True:
                if not prc_file in prc_files:
                    break
                else:
                    prc_file = "~" + prc_file

            shutil.copy(os.path.join(global_vars.project_folder, '.Исходники', source_file), os.path.join(global_vars.project_folder, '.Обработка', prc_file))

            self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                               f"Книга {source_file_number} из {len(source_files)}. Открываем для обработки {prc_file}")

            # проверяем возможно ли открыть файл и позволяют ли размер данных на листе сдвигать столбцы и строки
            wb = self.check_md_book(md_file, prc_file)

            if not wb:
                continue

            # обрабатываем листы
            for sheet_number, sheet in enumerate(wb.sheetnames, 1):

                ws = wb[sheet]

                ws_max_column = ws.max_column
                ws_max_row = ws.max_row

                # Делаем лист видимым
                ws.sheet_state = 'visible'

                # Записываем ширины колонок
                self.mysignal.emit(f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Записываем ширины колонок {prc_file} {sheet}")

                columns_width = []
                for i in range(1, ws_max_column + 1):
                    letter = get_column_letter(i) # преобразовываем индекс столбца в его букву
                    # получаем ширину столбца и добавляем в список
                    cw = ws.column_dimensions[letter].width
                    if cw:
                        columns_width.append(cw)
                    else:
                        columns_width.append(15)        

                # Записываем высоты строк
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Записываем высоты строк {prc_file} {sheet}")

                rows_height = []
                for i in range(1, ws_max_row + 1):
                    # получаем высоту столбца и добавляем в список
                    rh = ws.row_dimensions[i].height
                    if not rh:
                        rows_height.append(rh)                          
                    elif rh<10:                        
                        rows_height.append(10) 
                    elif rh>50:                        
                        rows_height.append(50)  
                    else:
                        rows_height.append(rh)

                # Снимаем пароль с листа
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Снимаем пароль с листа {prc_file} {sheet}")

                ws.protection.disable()

                # Показываем скрытые столбцы и строки
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Показываем скрытые столбцы и строки {prc_file} {sheet}")     
                    
                ws.column_dimensions.group(start='A', end=get_column_letter(ws_max_column), hidden=False)
                ws.row_dimensions.group(start=1, end=ws_max_row, hidden=False)

                # Удаляем проверку данных с листа
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Удаляем проверку данных с листа {prc_file} {sheet}")   

                ws.data_validations = DataValidationList()

                # Снимаем группировку колонок и столбцов
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Снимаем группировку колонок и столбцов {md_file} {sheet}")     

                ws.row_dimensions.group(1, ws_max_row, outline_level=0) # for entire sheet
                ws.column_dimensions.group('A', get_column_letter(ws_max_column), outline_level=0) # for entire sheet

                # Убираем фильтр
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. убираем фильтр {prc_file} {sheet}")   
                
                # ws.auto_filter.ref = None
                # ws.auto_filter.add_filter_column(None)
                ws.auto_filter.ref = ws.dimensions


                # Отменяем объединение ячеек
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Отменяем объединение ячеек {prc_file} {sheet}")   

                for merged_cell in list(ws.merged_cells.ranges):
                    ws.unmerge_cells(str(merged_cell))
 
                # Сдвигаем вниз
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Сдвигаем вниз {prc_file} {sheet}")   

                ws.insert_rows(idx=1, amount=2)
                
                # Сдвигаем вправо
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Сдвигаем вправо {prc_file} {sheet}")  

                ws.insert_cols(idx=1, amount=2)

                # Делаем ширины столбцов как в исходнике
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Сдвигаем вправо {prc_file} {sheet}")  

                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 5
                ws.column_dimensions[get_column_letter(ws_max_column + horizontal_offset + 1)].width = 5  

                for i, column_width in enumerate(columns_width, vertical_offset + 1):
                    letter = get_column_letter(i) 
                    ws.column_dimensions[letter].width = column_width

                ws.row_dimensions[1].height = 15 
                ws.row_dimensions[2].height = 15

                for i, row_height in enumerate(rows_height, vertical_offset + 1):
                    ws.row_dimensions[i].height = row_height                   


                # Очищаем от форматирования верхний ряд и левую колонку
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Сдвигаем вправо {prc_file} {sheet}")            
                for col in range(1, ws_max_column + 12):
                  
                    cell = ws.cell(column=col, row = 1)

                    cell.fill = no_fill
                    cell.border = no_border 
                    cell.alignment = styles.Alignment(wrap_text=True,vertical='center', horizontal='center')                  

                for row in range(1, ws_max_row + horizontal_offset + 1):
              
                    cell = ws.cell(column=1, row=row)
                    cell.fill = no_fill
                    cell.border = no_border 

                # Размечаем разделители
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Размечаем разделители {prc_file} {sheet}")   
            
                for col in range(1, ws_max_column + 12):
                    separator_cell = ws.cell(column=col, row = vertical_offset)
                    separator_cell.fill = sep_cell_style

              
                for col in range(1, ws_max_column + 4):
                    separator_cell = ws.cell(column=col, row = ws_max_row + vertical_offset + 1)
                    separator_cell.fill = sep_cell_style

                for row in range(1, ws_max_row + horizontal_offset + 1):
                    separator_cell = ws.cell(column=2, row = row)
                    separator_cell.fill = sep_cell_style

                
                for row in range(1, ws_max_row + horizontal_offset + 1):
                   
                    separator_cell = ws.cell(column=ws_max_column+3, row = row)
                    separator_cell.fill = sep_cell_style
                    
                # Замораживаем ячейки
                self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                   f"Книга {source_file_number} из {len(source_files)} лист {sheet_number} из {len(wb.sheetnames)}. Закрепляем диапазон {prc_file} {sheet}") 
                ws.sheet_view.topLeftCell = 'A1'
                freeze_cell = ws['C3']
                ws.freeze_panes = freeze_cell

                # Заполняем верхний заголовок
                # В ключе словаря первая цифра номер колонки, вторая номер строки
                # cell_khaki = styles.PatternFill(start_color='#c4bd97', fill_type='solid')
                cell_khaki = styles.PatternFill(start_color='C4BD97', fill_type='solid')
                cell_white = styles.PatternFill(start_color='FFFFFF', fill_type='solid')
                ws.row_dimensions[1].height = 45
                ws.row_dimensions[2].height = 45

                header = {
                    (1, 3): ('№', cell_khaki),
                    (1, 4): ('№ Вагона', cell_khaki),
                    (1, 5): ('Номер накладной', cell_khaki),
                    (1, 6): ('№ Контейнера', cell_khaki),
                    (1, 7): ('Ст. отправления', cell_khaki),
                    (1, 8): ('Ст. назначения', cell_khaki),
                    (1, 12): ('Дата отправки', cell_khaki),       
                    (1, 13): ('использование пути', cell_khaki),
                    (1, 14): (' тариф груженый', cell_khaki),
                    (2, ws_max_column+4): ('Отчет к акту выполненных работ №', cell_white),
                    (2, ws_max_column+5): ('Между', cell_white),
                    (2, ws_max_column+6): ('Дата составления', cell_white),
                    (2, ws_max_column+7): ('Наименование соисполнителя', cell_khaki),
                    (2, ws_max_column+8): ('№ договора с соисполнителем', cell_khaki),
                    (2, ws_max_column+9): ('Дата акта', cell_khaki),
                    (2, ws_max_column+10): ('Дата начала отчётного периода', cell_khaki),
                    (2, ws_max_column+11): ('Дата окончания отчётного периода', cell_khaki)
                    }

                for (y, x), (column_name, column_fill) in header.items():
                    header_cell = ws.cell(y, x)
                    header_cell.value = column_name
                    header_cell.fill = column_fill
                    header_cell.alignment = styles.Alignment(wrap_text=True, horizontal="center", vertical="center")


   

            # Сохраняем размеченную книгу.'
            self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                               f"Книга {source_file_number} из {len(source_files)}. Сохраняем книгу {prc_file}")

            wb.save(os.path.join(project_folder,'.Обработка', prc_file))
            wb.close()
            shutil.move(os.path.join(project_folder,'.Обработка', prc_file), os.path.join(project_folder,'.Размеченные', md_file))



    def all_columns(self, project_folder):
        global_vars.ui.info_label.setStyleSheet('color: blue')

        marked_folder = os.path.join(project_folder,'.Размеченные')

        #if os.path.exists(os.path.join(project_folder, "~$columns.xlsx")):
        #    self.warning_message ='Файл columns.xlsx уже открыт в Эксель.\nЗакройте его и заново нажмите кнопку "Обработка"'
        #    os.startfile(os.path.join(project_folder, "columns.xlsx")) 
        #    return
        
        files = [file for file in list(os.walk(os.path.join(project_folder, '.Размеченные')))[0][2] if file[0] != "~"]

        result_s_f_check_df = pd.DataFrame()

        result_first_and_second_line_df = pd.DataFrame()

        if files:
            for file_number, file in enumerate(files, 1):
                marked_file = os.path.join(marked_folder, file)
                with pd.ExcelFile(os.path.join(marked_folder,file)) as xlsx_file:
                    sheets = xlsx_file.sheet_names
                
                for sheet_number, sheet in enumerate(sheets, 1):
                    print(Fore.YELLOW, file, sheet, Fore.WHITE)                
                    self.mysignal.emit(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                    f"Книга {file_number} из {len(files)} лист {sheet_number} из {len(sheets)}. Считываем размеченные колонки из файла {file} из листа {sheet}") 

                    errors_list = []

                    sheet_df =  pd.read_excel(marked_file, sheet_name = sheet, dtype = str, header=None)
                    

                    if sheet_df.empty:
                        headers_df = pd.DataFrame([None, None])                    
                        s_f_check_dict = {'file':file, 'sheet':sheet, 's':'-', 'f':'-', 'Ошибки маркировки':'Пустой лист'}
                        # continue 

                    
                    else:
                        sheet_rem = (sheet_df[0].iloc[0])

                        errors_list.append(sheet_rem) # считываем комментарий если есть и добавляем в список
                        s = value_searcher(sheet_df[0], 's')
                        f = value_searcher(sheet_df[0], 'f')  
                        header_rows = sheet_df.iloc[0:2]
                        headers_df = sheet_df[sheet_df.columns[2:]].iloc[0:2] 

                        marking_errors = marking_checker(sheet_rem, s, f, header_rows)                    
                        s_f_check_dict = {'file':file, 'sheet':sheet, 's':s, 'f':f, 'Ошибки маркировки':marking_errors}              
                            
                    result_s_f_check_df = pd.concat([result_s_f_check_df, pd.DataFrame([s_f_check_dict])])

                    first_and_second_line_dict = {}

                    for column in headers_df.columns:
                        # print(headers_df)              
                        cell_in_first_line = headers_df.iloc[0].loc[column] 
                        cell_in_second_line = headers_df.iloc[1].loc[column]

                        if pd.notna(cell_in_first_line):
                            first_and_second_line_dict[cell_in_first_line] = cell_in_first_line
                        if pd.notna(cell_in_second_line):
                            first_and_second_line_dict[f"<<< с заполнением >>> {cell_in_second_line}"] = cell_in_second_line

                        first_and_second_line_df = pd.DataFrame([first_and_second_line_dict])


                    result_first_and_second_line_df = pd.concat([result_first_and_second_line_df, first_and_second_line_df]) 


                result_first_and_second_line_df.fillna("-", inplace=True)

            
            result_df = pd.concat([result_s_f_check_df, result_first_and_second_line_df], axis=1)  
            result_df_columns = list(dict.fromkeys(list(result_s_f_check_df.columns) +
                                                list(result_first_and_second_line_df.columns)))
            
            result_df = result_df[result_df_columns]        


            errors_list = list(map(str, errors_list))


            result_df.to_excel(os.path.join(project_folder, "columns.xlsx"), index=False)   
            
            wb = load_workbook(os.path.join(project_folder, "columns.xlsx"))
            ws = wb.active

            # Ширину столбцов A, B задаём по содержимому
            column_a = ws['A']
            max_a = 0
            for i in column_a:
                max_a=max(max_a, len(str(i.value)))

            ws.column_dimensions["A"].width = max_a+2


            column_b = ws['B']
            max_b = 0
            for i in column_b:
                max_b=max(max_b, len(str(i.value)))                

            ws.column_dimensions["B"].width = max_b+2


            global_vars.ui.pushButtonConcat.setEnabled(True)
            global_vars.ui.pushButtonMakeFiles.setEnabled(True)             

            # Закрепляем области
            freeze_cell = ws.cell(column=3, row=2)
            ws.freeze_panes = freeze_cell

            ws.auto_filter.ref = ws.dimensions

            wb.save(os.path.join(project_folder, "columns.xlsx"))
            global_vars.ui.info_label.setStyleSheet('color: green')  


    def on_signal(self,mysignal):          
        global_vars.ui.info_label.setText(mysignal)


    def run(self): 
        self.message_title = "Обработка"
        self.error_message = ""
        self.warning_message = ""

        if os.path.exists(os.path.join(global_vars.project_folder, "~$columns.xlsx")):
            global_vars.ui.info_label.setStyleSheet('color: red')             
            global_vars.ui.info_label.setText('Закройте файл columns.xlsx перед тем как запустить обработку.')   
            self.warning_message ='Файл columns.xlsx уже открыт на рабочем столе.\nЗакройте его и заново нажмите кнопку "Обработка"'
            return 
        else:
            df = pd.DataFrame(['Что-то пошло не так'], index=None)
            df.to_excel(os.path.join(global_vars.project_folder, 'columns.xlsx'), index=None, header=None) 
           
        self.err_list = [] 

        clean_process_folder(global_vars.project_folder)  
        self.clean_md_folder()
        
        is_files_modified = check_files_modified('.Исходники')
        if type(is_files_modified) == type([]):
            print('Мы тут')

            # если файлы в .Исходниках поменялись и они есть в .Размеченных
            # то нужно спросить нужно ли его переразметить или оставить как есть
            md_files = list(os.walk(os.path.join(global_vars.project_folder, '.Размеченные')))[0][2] 
            print(Fore.GREEN, is_files_modified, Fore.RESET)
            print(Fore.GREEN, md_files, Fore.RESET)
            self.wrn_list = []
            for file_modified in is_files_modified:
                print(Fore.CYAN, file_modified, file_modified in is_files_modified, Fore.RESET)
                if f"md_{file_modified}" in md_files:
                    self.warning_message = ('Некоторые файлы в папке .Исходники\n'
                                            'были пересохранены.\n'
                                            'Если их нужно переразметить,\n'
                                            'удалите md-файлы из папки .Размеченные!')            
                    self.wrn_list.append((f"md_{file_modified}", 'Файл в .Исходниках поменялся, если его нужно переразметить, удалите md-файл из папки .Размеченные'))

            if self.wrn_list:
                df = pd.DataFrame(self.wrn_list)
                df.to_excel(os.path.join(global_vars.project_folder, 'columns.xlsx'), index=None, header=None)
                
            refresh_files_info('.Исходники')
            refresh_files_info('.Размеченные')


        if not self.warning_message:
            self.check_src_files_available()
        else:
            if self.err_list:
                df = pd.DataFrame(self.err_list)
                df.to_excel(os.path.join(global_vars.project_folder, 'columns.xlsx'), index=None, header=None)

        if not self.error_message and not self.warning_message:
            self.premarker(global_vars.project_folder)
        else:
            if self.err_list:
                df = pd.DataFrame(self.err_list)
                df.to_excel(os.path.join(global_vars.project_folder, 'columns.xlsx'), index=None, header=None)

        if not self.error_message and not self.warning_message:
            self.all_columns(global_vars.project_folder)
        else:
            if self.err_list:
                df = pd.DataFrame(self.err_list)
                df.to_excel(os.path.join(global_vars.project_folder, 'columns.xlsx'), index=None, header=None)

                  


    def on_clicked(self):
        init_project()
              
        self.start() # Запускаем поток  
     


    def on_started(self): # Вызывается при запуске потока
        global_vars.ui.pushButtonChooseProjectFolder.setEnabled(False)
        global_vars.ui.pushButtonXLStoXLSX.setEnabled(False)
        global_vars.ui.pushButtonProcessing.setEnabled(False)
        global_vars.ui.pushButtonOpenChoosedFiles.setEnabled(False)
        global_vars.ui.pushButtonOpenChoosedMDFiles.setEnabled(False)
        global_vars.ui.pushButtonDelChoosedMDFiles.setEnabled(False)
        global_vars.ui.pushButtonConcat.setEnabled(False)
        global_vars.ui.pushButtonMakeFiles.setEnabled(False)
        global_vars.ui.info_label.setStyleSheet('color: blue')


    def on_finished(self): # Вызывается при завершении потока
        global_vars.ui.pushButtonChooseProjectFolder.setEnabled(True)
        global_vars.ui.pushButtonXLStoXLSX.setEnabled(True)
        global_vars.ui.pushButtonProcessing.setEnabled(True)
        global_vars.ui.pushButtonOpenChoosedFiles.setEnabled(True)
        global_vars.ui.pushButtonDelChoosedMDFiles.setEnabled(True)
        global_vars.ui.pushButtonOpenChoosedMDFiles.setEnabled(True)


        if self.error_message:
            global_vars.ui.info_label.setStyleSheet('color: red')
            global_vars.ui.info_label.setText(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                              f"{self.error_message.replace('\n',' ')}")
            QtWidgets.QMessageBox.critical(None,
                                           self.message_title,
                                           self.error_message,
                                           buttons=QtWidgets.QMessageBox.StandardButton.Ok)
            refresh_files_info('.Исходники')
            refresh_files_info('.Размеченные')

        elif self.warning_message:
            # print(Fore.RED, self.err_list, Fore.RED)
            global_vars.ui.info_label.setStyleSheet('color: red')
            global_vars.ui.info_label.setText(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                              f"{self.warning_message.replace('\n',' ')}")
            
            QtWidgets.QMessageBox.warning(None,
                                           self.message_title,
                                           self.warning_message,
                                           buttons=QtWidgets.QMessageBox.StandardButton.Ok)
        else:
            global_vars.ui.info_label.setStyleSheet('color: green')
            global_vars.ui.info_label.setText(f"{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")} "
                                              f"Обработка завершена. Открываем columns.xlsx на рабочем столе.")
            refresh_files_info('.Исходники')
            refresh_files_info('.Размеченные')


        # if os.path.exists(os.path.join(global_vars.project_folder, "columns.xlsx")):
        os.startfile(os.path.join(global_vars.project_folder, "columns.xlsx"))
